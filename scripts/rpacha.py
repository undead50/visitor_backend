
def selenium_script(url):
    try:
        from selenium import webdriver
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.common.by import By
        import openpyxl
        import time
        import os


        ###########################################################################################
        ## INITIALIZE DRIVER
        driver = webdriver.Chrome()


        ############################################################################################
        ## XLSX DATA RETRIEVAL PROCESS
        file_name = "challenge.xlsx"
        path = "C:\\Users\\AYUSH.PRADHAN\\Documents\\GitHub\\rpabot\\xlxs"
        file_path  = os.path.join(path, file_name)
        
        wb_obj = openpyxl.load_workbook(file_path)
        sheet_obj = wb_obj.active

        # Shrinking the list to 7 elements at a time to represent a single users information. 
        def list_creation(xlsx_data_list):
            input_list = [] 
            idx = 0
            for i in range(row_count -1): # 10
                for j in range(max_col- 1): # 7
                    if idx == len(xlsx_data_list):
                        break
                    else:
                        input_list.append(xlsx_data_list[idx])
                        idx += 1
                data_input(input_list)

        # Count of the Rows with data on XLSX File
        row_count = 0
        for row in sheet_obj:
            if not all([cell.value == None for cell in row]):
                row_count += 1

        # Count of the Columns on XLSX File
        max_col = sheet_obj.max_column

        # Creating a big list with all data values from XLSX File. 
        # Going cell by cell, left to right, each row at a time.
        xlsx_data_list = []
        for i in range(2, row_count + 1):
            for j in range(1, max_col):
                cell_obj = sheet_obj.cell(row = i, column= j)
                xlsx_data_list.append(cell_obj.value)

        
     


        ############################################################################################
        ## SELENIUM PROCESS

        chrome_driver_path = "location of chromedriver.exe on computer"
        
        driver.maximize_window()


        # Inputting the data onto the RPA Website 
        def data_input(input_list):
            # "Getting" the start button on the RPA Website 
            start_button = driver.find_element(By.CLASS_NAME,'btn-large.uiColorButton')
            count = 0
            while count < 1:
                start_button.click()
                count += 1
            # "Getting" multiple form HTML elements on the RPA Website 
            first_name_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelFirstName"]')
            last_name_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelLastName"]')
            cmpny_name_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelCompanyName"]')
            role_in_cmpny_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelRole"]')
            address_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelAddress"]')
            email_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelEmail"]')
            phone_num_form = driver.find_element(By.XPATH,'//*[@ng-reflect-name="labelPhone"]')
            submit_button = driver.find_element(By.CLASS_NAME,'btn.uiColorButton')
            # Inputting the Data
            first_name_form.click()
            first_name_form.send_keys(input_list[0]) # 0 = First Name 
            last_name_form.click()
            last_name_form.send_keys(input_list[1]) # 1 = Last Name 
            cmpny_name_form.click()
            cmpny_name_form.send_keys(input_list[2]) # 2 = Company Name 
            role_in_cmpny_form.click()
            role_in_cmpny_form.send_keys(input_list[3]) # 3 = Role in Company 
            address_form.click()
            address_form.send_keys(input_list[4]) # 4 = Address
            email_form.click()
            email_form.send_keys(input_list[5]) # 5 = Email
            phone_num_form.click()
            phone_num_form.send_keys(input_list[6]) # 6 = Phone Number 
            submit_button.click()
            input_list.clear() # Clears the current data elements in the list
            return

        # Open RPA website #
        rpa_website = driver.get("https://www.rpachallenge.com/")
        list_creation(xlsx_data_list)
        time.sleep(5)
        
    except Exception as e:
        print(e)

    finally:
        driver.quit()        

   
############################################################################################